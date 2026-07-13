# -*- coding: utf-8 -*-
"""
Created on Thu Jun 25 13:51:06 2026

@author: Dennis Corraliza

Error Analysis
"""

import numpy as np
import geometry
import examples
import assembly_vec as assembly
import os
from scipy.spatial import Delaunay
import openpyxl

def delaunay_lumped_weights(points, tri):
    """
    Lumped integration weights for a 2D point cloud.

    Parameters
    ----------
    points : (N,2) ndarray
    tri : scipy.spatial.Delaunay
        Delaunay triangulation.

    Returns
    -------
    weights : (N,) ndarray
        Nodal integration weights.
    """

    weights = np.zeros(len(points))

    for simplex in tri.simplices:
        p0, p1, p2 = points[simplex]

        v1 = p1 - p0
        v2 = p2 - p0
        
        area = 0.5 * abs(v1[0] * v2[1] - v1[1] * v2[0])
        weights[simplex] += area / 3.0

    return weights

def l2_error(u_soln, u_ex):
    error = abs(u_soln-u_ex)
    return np.linalg.norm(error)/np.sqrt(len(error))

def l2_error_relative(u_soln, u_ex):
    error = abs(u_soln-u_ex)
    return np.linalg.norm(error)/np.linalg.norm(u_ex)   

def max_error(u_soln, u_ex):
    return np.max(abs(u_soln-u_ex))

def max_error_relative(u_soln, u_ex):
    return np.max(abs(u_soln-u_ex))/np.max(u_ex)

def energy_error(context, weights, u_soln, u_ex, sparse=False):
    P = context.nodes
    A = context.A
    num_nodes = len(P)
    dim = len(P[0])

    if sparse:
        W_grads = assembly.global_grads_sparse(context)
    else:
        W_grads = assembly.global_grads(context)
        
    err = u_soln-u_ex
    
    grad_err = np.zeros((num_nodes, dim))
    energy_err = 0.0
    
    for i in range(dim):
        if sparse:
            grad_err[:,i] = W_grads[i] @ err
        else:
            grad_err[:,i] = W_grads[:,:,i] @ err
    
    for i in range(num_nodes):
        energy_err += weights[i] * np.dot(grad_err[i,:], A @ grad_err[i,:])
    
    return np.sqrt(energy_err) 

def energy_error_delaunay(context, u_soln, u_ex, sparse=False):
    P = context.nodes
    tri = Delaunay(P)
    weights = delaunay_lumped_weights(P, tri)
    #print(weights.sum())
    
    return energy_error(context, weights, u_soln, u_ex, sparse)

def energy_error_delaunay_relative(context, u_soln, u_ex, sparse=False):
    P = context.nodes
    tri = Delaunay(P)
    weights = delaunay_lumped_weights(P, tri)
    
    energy_err = energy_error(context, weights, u_soln, u_ex, sparse)
    dirichlet_energy = energy_error(context, weights, u_ex,
                                    np.zeros(u_ex.shape), sparse)
    
    return energy_err/dirichlet_energy

def energy_functional(context, weights, u_soln, f_vec, sparse=False):
    P = context.nodes
    A = context.A
    num_nodes = len(P)
    dim = len(P[0])

    if sparse:
        W_grads = assembly.global_grads_sparse(context)
    else:
        W_grads = assembly.global_grads(context)
    
    grad_u = np.zeros((num_nodes, dim))
    energy= 0.0
    
    for i in range(dim):
        if sparse:
            grad_u[:,i] = W_grads[i] @ u_soln
        else:
            grad_u[:,i] = W_grads[:,:,i] @ u_soln
    
    for i in range(num_nodes):
        energy += weights[i]*(np.dot(grad_u[i,:], A @ grad_u[i,:])/2
                              - f_vec[i] * u_soln[i])
    
    return energy

def energy_functional_delaunay(context, u_soln, f_vec, sparse=False):
    P = context.nodes
    tri = Delaunay(P)
    weights = delaunay_lumped_weights(P, tri)
    
    return energy_functional(context, weights, u_soln, f_vec, sparse)

def coeff_matrix(eig_1, eig_2, rad24):
    """
    Build a symmetric 2x2 anisotropic diffusion tensor from eigenvalues
    and a rotation angle.

    Constructs a diagonal matrix `D = diag(eig1, eig2)` representing the
    diffusion coefficients along the tensor's principal axes, then
    rotates it by an angle `theta = pi * rad24 / 24` (i.e. `rad24` is
    expressed in 24ths of a half-turn, or 7.5-degree increments) via
    `A = V @ D @ V.T`, where `V` is the 2D rotation matrix for `theta`.
    The result is the diffusion tensor `A` expressed in the original
    (unrotated) x-y coordinate frame.

    Parameters
    ----------
    eig_1 : float
        Diffusion coefficient (eigenvalue) along the tensor's first
        principal axis.
    eig_2 : float
        Diffusion coefficient (eigenvalue) along the tensor's second
        principal axis.
    rad24 : float
        Rotation angle of the principal axes, in units of pi/24
        radians (e.g. `rad24 = 1` corresponds to a 7.5-degree
        rotation, `rad24 = 12` corresponds to 90 degrees).

    Returns
    -------
    numpy.ndarray, shape (2, 2)
        Symmetric positive (semi-)definite diffusion tensor `A`,
        rotated from the principal-axis frame into the standard x-y
        frame.
    """
    
    theta = np.pi*rad24/24.0
    D = np.array([[eig_1, 0.0], [0.0, eig_2]])
    V = np.array([[np.cos(theta), -np.sin(theta)],
                  [np.sin(theta), np.cos(theta)]])
    
    return V @ D @ V.T  

def pde_context_provider(N_int, eig_1, eig_2, num_stencil_nodes,
                         num_rings, rbf_shape, eps, augmentation,
                         rad24, example_problem, sparse=False):
    
    # Define the parameters of the radial basis function
    tol = 1e-12

    # -----------------------------
    # BUILD NODES
    # -----------------------------
    Nx = N_int
    Ny = N_int
    L = 1.0
    shape = 'square'
    
    P, num_int = geometry.uniform_int_square(L, Nx, Ny)
        
    # -----------------------------
    # ANISOTROPY AND PDE PROPERTIES
    # -----------------------------    
    # Define the conductivity condition
    A = coeff_matrix(eig_1, eig_2, rad24)
    print("Coefficient Matrix:\n" + str(A))
    
    # Forcing term parameters
    Amp = 1e3
    modes = [1.0,3.0]
    
    # -----------------------------
    # BUILD TEST CASE AND SOLVE
    # -----------------------------
    f, g, btype, u_exact = example_problem(Amp, modes, A)
    
    context = assembly.rbf_fd_system(f, g, btype, P,
                                 rbf_shape, shape, L, num_stencil_nodes,
                                 num_rings, augmentation=augmentation,
                                 A=A, eps=eps, tol=tol, sparse=sparse)
    
    P = context.nodes
    W = context.W
    F = context.F
    
    if sparse:
        u_soln = assembly.rbf_fd_solve_sparse(W, F)
    else:
        u_soln = assembly.rbf_fd_solve(W, F)
        print(f"Condition:    {np.linalg.cond(W): e}")
        
    u_ex = u_exact(P.T)    
    return context, u_soln, u_ex

def error_analysis(Nx, Ny, num_stencil_nodes, num_rings, eig_1,
                   eig_2, rbf_shape, eps, augmentation, rad24,
                   context, u_soln, u_ex, sparse=False):
    W = context.W
    F = context.F
    
    print(f'Sparse Solve: {sparse}')
    print(f'Nx = {Nx}, Ny = {Ny}')
    print(f'Number of Stencils Nodes = {num_stencil_nodes}')
    print(f'Number of Rings          = {num_rings}')
    print(f'RBF: {rbf_shape} with augmentation: {augmentation}')
    
    Lu_approx = W @ u_ex
    
    res_max    = max_error_relative(Lu_approx, F)
    res_l2     = l2_error_relative(Lu_approx, F)
    res_energy = energy_error_delaunay_relative(context, u_soln, u_ex, sparse)
    print("Max L_h u_exact-f Error Rel = ", res_max)
    print("L2  L_h u_exact-f Error Rel = ", res_l2)
    print("Energy Error Rel            = ", res_energy)
    print()
    
    return {
    'Nx': Nx,
    'Ny': Ny,
    'eps': eps,
    'num_stencil_nodes': num_stencil_nodes,
    'num_rings': num_rings,
    'eigenvalue 1': eig_1,
    'eigenvalue 2': eig_2,
    'radian = r/24': rad24,
    'augmentation': augmentation,
    'sparse': sparse,
    'max_error_rel': res_max,
    'l2_error_rel': res_l2,
    'energy_error_rel': res_energy,
    }

# -----------------------------
# WRITE TO EXCEL
# -----------------------------
def append_sheet_to_excel(sheet_name, rows, filepath):
    if os.path.exists(filepath):
        wb = openpyxl.load_workbook(filepath)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_is_new = False
    else:
        ws = wb.create_sheet(title=sheet_name[:31])
        sheet_is_new = True

    if not rows:
        wb.save(filepath)
        return

    headers = list(rows[0].keys())

    if sheet_is_new or ws.max_row == 1 and ws.cell(1, 1).value is None:
        ws.append(headers)

    for row in rows:
        ws.append([row[h] for h in headers])

    wb.save(filepath)
    print(f"Saved results to {filepath}")

def data_output(example_num):
    example = eval('examples.example_'+str(example_num))

    script_dir = os.path.dirname(os.path.abspath(__file__))
    parent_dir = os.path.dirname(script_dir)
    results_dir = os.path.join(parent_dir, 'Results')
    os.makedirs(results_dir, exist_ok=True)
    
    output_path = os.path.join(results_dir, 
                               f'rbf_fd_parameter_study_{example_num}.xlsx')

    # -----------------------------
    # PRACTICE RUN: CHECK COMPILE
    # -----------------------------
    print('=======================Code Compiling Study=======================')
    eps = 3.0
    rbf_shape = 'gaussian'
    augmentation = False
    
    # Provide whether the solve is sparse or dense
    sparse = True
    
    N_int = 30
    eig_1 = 1e0
    eig_2 = 1e0
    rad24 = 0.0    
    num_stencil_nodes = 25
    num_rings = 5
    
    context, u_soln, u_ex = pde_context_provider(N_int, eig_1, eig_2,
                                                 num_stencil_nodes,
                                                 num_rings, rbf_shape, eps,
                                                 augmentation, rad24,
                                                 example, sparse)
    error_analysis(N_int, N_int, num_stencil_nodes, num_rings, eig_1,
                   eig_2, rbf_shape, augmentation, rad24, context,
                   u_soln, u_ex, sparse)
    
    # -----------------------------
    # BASELINE PARAMETERS FOR EACH STUDY
    # (mirrors the values each TEST block set before its loop in the
    # original sequential script)
    # -----------------------------
    N_ints = [20, 30, 50, 75, 100, 125, 150]
    INV_L_S = [5e-2,1e-1,5e-1,1.0,2.0,2.5,3.0,3.5,4.0,5.0,7.5,10.0,15.0]
    N_S_N = [15, 25, 40, 50, 65, 75, 85, 100, 115, 130, 150, 200]
    N_C_R = [5, 6, 7, 8, 9, 10, 11, 12]
    Eig_R_2 = [1e1,5e0,1e0,5e-1,1e-1,5e-2,1e-2,5e-3,1e-3,5e-4,1e-4,5e-5,1e-5]
    Eig_RAD_24 = [0.0, 4.0, 6.0, 8.0, 12.0, 16.0, 18.0, 20.0, 24.0]
    
    N_int = 100
    eig_1 = 1e0
    eig_2 = 5e-3
    rad24 = 12.0    
    num_stencil_nodes = 100
    num_rings = 10

    # -----------------------------
    # TEST 1: INTERIOR GRID SIZE
    # -----------------------------
    print('=================1: Interior Grid Size Study======================')
    #N_ints = [30]
    rows = []
    
    for x in N_ints:
        print(f'---------------------N_int = {x}-------------------------')
        context, u_soln, u_ex = pde_context_provider(x, eig_1, eig_2,
                                                     num_stencil_nodes,
                                                     num_rings, rbf_shape,
                                                     eps, augmentation,
                                                     rad24, example, sparse)
        row = error_analysis(x, x, num_stencil_nodes, num_rings, eig_1,
                             eig_2, rbf_shape, eps, augmentation, rad24, context,
                             u_soln, u_ex, sparse)
        row['varied_param'] = 'N_int'
        row['varied_value'] = x
        rows.append(row)
    append_sheet_to_excel('Grid Size', rows, output_path)

    # -----------------------------
    # TEST 2: INVERSE LENGTH SCALE
    # -----------------------------
    print('=================1: Interior Grid Size Study======================')
    #INV_L_S = [3.0]
    rows = []
    
    for x in INV_L_S:
        print(f'---------------Inverse Length Scale = {x}--------------------')
        context, u_soln, u_ex = pde_context_provider(N_int, eig_1, eig_2,
                                                     num_stencil_nodes,
                                                     num_rings, rbf_shape,
                                                     x, augmentation,
                                                     rad24, example, sparse)
        row = error_analysis(N_int, N_int, num_stencil_nodes, num_rings, eig_1,
                             eig_2, rbf_shape, x, augmentation, rad24, context,
                             u_soln, u_ex, sparse)
        row['varied_param'] = 'Inverse Length Scale'
        row['varied_value'] = x
        rows.append(row)
    append_sheet_to_excel('Grid Size', rows, output_path)    

    # -----------------------------
    # TEST 2: NUMBER STENCIL NODES
    # -----------------------------
    print('================2: Number of Stencil Nodes Study==================')
    #N_S_N = [25]   
    rows = []
    
    for x in N_S_N:
        print(f'----------num_stencil_nodes = {x}------------')
        context, u_soln, u_ex = pde_context_provider(N_int, eig_1, eig_2,
                                                     x,
                                                     num_rings, rbf_shape,
                                                     eps, augmentation,
                                                     rad24, example, sparse)
        row = error_analysis(N_int, N_int, x, num_rings, eig_1,
                             eig_2, rbf_shape, eps, augmentation, rad24, context,
                             u_soln, u_ex, sparse)
        row['varied_param'] = 'num_stencil_nodes'
        row['varied_value'] = x
        rows.append(row)
    append_sheet_to_excel('Stencil Nodes', rows, output_path)
    
    # -----------------------------
    # TEST 3: NUMBER CENTER RINGS
    # -----------------------------
    print('=================3: Number of Center Rings Study==================')
    #N_C_R = [5]
    rows = []
    
    for x in N_C_R:
        print(f'-----------------num_rings = {x}---------------------')
        context, u_soln, u_ex = pde_context_provider(N_int, eig_1, eig_2,
                                                     num_stencil_nodes,
                                                     x, rbf_shape,
                                                     eps, augmentation,
                                                     rad24, example, sparse)
        row = error_analysis(N_int, N_int, num_stencil_nodes, x, eig_1,
                             eig_2, rbf_shape, eps, augmentation, rad24, context,
                             u_soln, u_ex, sparse)
        row['varied_param'] = 'num_rings'
        row['varied_value'] = x
        rows.append(row)
    append_sheet_to_excel('Center Rings', rows, output_path)
    
    # -----------------------------
    # TEST 4: EIGENVALUE RATIO
    # -----------------------------
    print('====================4: Eigenvalue Ratio Study=====================')
    #Eig_R_2 = [1e0]
    rows = []
    
    for x in Eig_R_2:
        print(f'--------------------eig_2 = {x}--------------------------')
        context, u_soln, u_ex = pde_context_provider(N_int, eig_1, x,
                                                     num_stencil_nodes,
                                                     num_rings, rbf_shape,
                                                     eps, augmentation,
                                                     rad24, example, sparse)
        row = error_analysis(N_int, N_int, num_stencil_nodes, num_rings, eig_1,
                             x, rbf_shape, eps, augmentation, rad24, context,
                             u_soln, u_ex, sparse)
        row['varied_param'] = 'eig_2'
        row['varied_value'] = x
        rows.append(row)
    append_sheet_to_excel('Eigenvalue Ratio', rows, output_path)

    # -----------------------------
    # TEST 5: EIGENVECTOR ANGLE
    # -----------------------------
    print('===================5: Eigenvector Radian Study====================')
    #Eig_RAD_24 = [0.0] 
    rows = []
    
    for x in Eig_RAD_24:
        print(f'-----------------radian=rad24/24 = {x}-------------------')
        context, u_soln, u_ex = pde_context_provider(N_int, eig_1, eig_2,
                                                     num_stencil_nodes,
                                                     num_rings, rbf_shape,
                                                     eps, augmentation,
                                                     x, example, sparse)
        row = error_analysis(N_int, N_int, num_stencil_nodes, num_rings, eig_1,
                             eig_2, rbf_shape, eps, augmentation, x, context,
                             u_soln, u_ex, sparse)
        row['varied_param'] = 'rad24'
        row['varied_value'] = x
        rows.append(row)
    append_sheet_to_excel('Eigenvector Angle', rows, output_path)

if __name__ == "__main__":
    example_nums = [2, 3, 4, 5, 6, 7, 8, 9, 10]
    
    for example_num in example_nums:
        data_output(example_num)